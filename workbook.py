from openpyxl import Workbook, load_workbook


print("Writing 1,000,000 rows in standard mode...")
wb_standard = Workbook()
ws_standard = wb_standard.active

for i in range(1, 1000001):
    ws_standard.append([f"Row {i}", f"Value {i}"])

standard_file = "standard_mode.xlsx"
wb_standard.save(standard_file)
print(f"Saved standard Excel file as '{standard_file}'")


print("\nReading first 10 rows from the standard Excel file...")
wb_read = load_workbook(standard_file, read_only=True)
ws_read = wb_read.active

for row in ws_read.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)


print("\nEfficiently writing 1,000,000 rows using write_only mode...")
wb_fast = Workbook(write_only=True)
ws_fast = wb_fast.create_sheet()

for i in range(1, 1000001):
    ws_fast.append([f"Row {i}", f"Data {i}"])

fast_file = "write_only_huge_data.xlsx"
wb_fast.save(fast_file)
print(f"Saved optimized Excel file as '{fast_file}'")


print("\nReading filtered data (first 10 rows and 5 columns)...")
wb_filtered = load_workbook(fast_file, read_only=True)
ws_filtered = wb_filtered.active

for row in ws_filtered.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
    print(row)
